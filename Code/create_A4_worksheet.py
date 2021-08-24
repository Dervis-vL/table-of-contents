import openpyxl, os
from openpyxl.styles import Font, Alignment

def create_A4_TOC(dir, ChapAndPar):
    '''
    This function creates a file in the same directory as the input file named
    Table_of_content.xlsx. 
    Link this finction to a different fuction to move it to the recommended file.

    Before using this file, check if it already exists a file with the exact same name.
    For questions: dervis.vanleersum@gmail.com
    '''

    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    # the created file containing the worksheet with the table of content will get this name:
    fileName = 'Table_of_content.xlsx'

    # check directory for doubles
    duplicate = 0
    for file in os.listdir(dir):
        if fileName in file:
            duplicate += 1
    
    pathTOC = dir + "\\" + str(duplicate) + "_" + fileName

    # create new workbook and worksheet named 'TOC'
    title = 'TOC'

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = title

    # format page layout height and width
    # height conversion 1 part is 1/28.36 cm; 1 px = 0.75
    # height conversion 1 part is 1/72 inch
    for i in range(1,4400):
        ws.row_dimensions[i].height = 14.25 #19 px
    for j in alphabet:
        ws.column_dimensions[j].width = 7.1 #54 px

    # Create table of content header
    ws['A1'].value = 'Inhoudsopgave'

    # Font format
    # Iv groen = RGB(0, 115, 97); HEX(00007361)
    ws['A1'].font = Font(name='Arial',
                            sz=18,
                            color='00007361')

    # change alignment
    ws['A1'].alignment = Alignment(horizontal="left", vertical="top")

    # merge cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=12)

    # page orientation
    ws.set_printer_settings(paper_size=9, orientation='portrait')

    # page print setup (input in inches)
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.71
    ws.page_margins.top = 1.18
    ws.page_margins.bottom = 1.18
    ws.page_margins.header = 0.08
    ws.page_margins.footer = 0.08

    # page orientation; This is not linked to the rest of the report. For easy printability it's all down
    #ws.page.PrintPageSetup()
    ws.page_setup.pageOrder = 'overThenDown'

    # create header w/ page number or footer w/ text
    # ws.oddHeader.right.text = "Test bitch"
    # ws.oddHeader.right.size = 14
    # ws.oddHeader.right.font = 'Tahoma,Bold'
    # ws.oddHeader.right.color = 'CC3366'
    ws.oddFooter.right.text = "&[Page]/&N\n\n"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.font = 'Arial'

    ws.sheet_view.view = 'pageLayout'

    # Get amount chapters and paragraphs; add total number
    keys = 4
    pars = 0
    for key, value in ChapAndPar.items():
        keys += 1
        for par in value:
            pars += 1
        pars += 1
    pars -= 1

    # create print area based on the number of lines the TOC has
    sums = keys + pars
    pages = (sums // 48) + 1
    rowsTOC = 48 * pages
    ws.print_area = 'A1:L' + str(rowsTOC)
    
    # print chapters and paragraphs to the file
    line = 5
    for key, value in ChapAndPar.items():
        ws.cell(line, 1).value = key.split(" ", 1)[0]
        ws.cell(line, 1).font = Font(name='Arial',
                                    sz=12,
                                    color='00000000',
                                    bold=True)
        ws.cell(line, 1).alignment = Alignment(horizontal="left", vertical="bottom")

        ws.cell(line, 2).value = key.split(" ", 1)[1].lstrip()
        ws.cell(line, 2).font = Font(name='Arial',
                                    sz=12,
                                    color='00000000',
                                    bold=True)
        ws.cell(line, 2).alignment = Alignment(horizontal="left", vertical="bottom")
        line += 1

        for par in value:
            ws.cell(line, 1).value = par.split(" ", 1)[0]
            ws.cell(line, 1).font = Font(name='Arial',
                                        sz=10,
                                        color='00000000',
                                        bold=True)
            ws.cell(line, 1).alignment = Alignment(horizontal="left", vertical="bottom")

            ws.cell(line, 2).value = par.split(" ", 1)[1].lstrip()
            ws.cell(line, 2).font = Font(name='Arial',
                                        sz=10,
                                        color='00000000',
                                        bold=False)
            ws.cell(line, 2).alignment = Alignment(horizontal="left", vertical="bottom")
            line += 1
        ws.cell(line, 1).value = " "
        line += 1

    # close and save file with full table of content
    wb.save(pathTOC)
    wb.close()

    return pathTOC, pages


        # last edited: 27-05-2021
def create_A4_TOC_empty(dir, ChapAndPar):
    '''
    This function creates an empty worksheet in the main report to simulate the amount of pages the TOC will be,
    in order to export the .xlsx file to pdf with accurate page numbers.
    For questions: dervis.vanleersum@gmail.com
    '''

    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    # the created file containing the worksheet with the table of content will get this name:
    fileName = 'Table_of_content.xlsx'

    # check directory for doubles
    duplicate = 0
    for file in os.listdir(dir):
        if fileName in file:
            duplicate += 1
    
    pathTOC = dir + "\\" + str(duplicate) + "_" + fileName

    # create new workbook and worksheet named 'TOC'
    title = 'TOC'

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = title
    ws.sheet_properties.tabColor = "007161"

    # format page layout height and width
    # height conversion 1 part is 1/28.36 cm; 1 px = 0.75
    # height conversion 1 part is 1/72 inch
    for i in range(1,4400):
        ws.row_dimensions[i].height = 14.25 #19 px
    for j in alphabet:
        ws.column_dimensions[j].width = 7.1 #54 px

    # Create table of content header
    ws['A1'].value = '__%REPLACE%__'

    # Font format
    # Iv groen = RGB(0, 115, 97); HEX(00007361)
    ws['A1'].font = Font(name='Arial',
                            sz=18,
                            color='00007361')

    # change alignment
    ws['A1'].alignment = Alignment(horizontal="left", vertical="top")

    # merge cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=12)

    # page orientation
    ws.set_printer_settings(paper_size=9, orientation='portrait')

    # page print setup (input in inches)
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.71
    ws.page_margins.top = 1.18
    ws.page_margins.bottom = 1.18
    ws.page_margins.header = 0.08
    ws.page_margins.footer = 0.08

    # page orientation; This is not linked to the rest of the report. For easy printability it's all down
    #ws.page.PrintPageSetup()
    ws.page_setup.pageOrder = 'overThenDown'

    # create header w/ page number or footer w/ text
    # ws.oddHeader.right.text = "Test bitch"
    # ws.oddHeader.right.size = 14
    # ws.oddHeader.right.font = 'Tahoma,Bold'
    # ws.oddHeader.right.color = 'CC3366'
    ws.oddFooter.right.text = "&[Page]/&N\n\n"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.font = 'Arial'

    ws.sheet_view.view = 'pageLayout'

    # Get amount chapters and paragraphs; add total number
    keys = 4
    pars = 0
    for key, value in ChapAndPar.items():
        keys += 1
        for par in value:
            pars += 1
        pars += 1
    pars -= 1

    # create print area based on the number of lines the TOC has
    sums = keys + pars
    pages = (sums // 48) + 1
    rowsTOC = 48 * pages
    ws.print_area = 'A1:L' + str(rowsTOC)

    # close and save file with full table of content
    wb.save(pathTOC)
    wb.close()

    return pathTOC, pages, title



# # TODO create an Excel front page in every style
# def create_A4_FP():
#     # TODO create input


# # TODO create an Excel report format; combine this with 'FP' and 'TOC'
# def create_A4_Report():
#     # TODO create input


# # TODO create a company colofon; combine this with 'FP', 'TOC' and 'Report'
# def create_A4_COL():
#     # TODO create input

# TODO move worksheets between files


        # last edited: 27-05-2021
def update_A4_TOC(path, ChapAndPar, pages):
    '''
    This function updates the Table_of_content.xlsx file
    Link this finction to a different fuction to move it to the recommended file.

    Before using this file, check if it already exists a file with the exact same name.
    For questions: dervis.vanleersum@gmail.com
    '''

    # update the worksheet named 'TOC'
    title = 'TOC'

    wb = openpyxl.load_workbook(path)
    ws = wb[title]

    # Create table of content header
    ws['A1'].value = 'Inhoudsopgave'

    # Font format
    # Iv groen = RGB(0, 115, 97); HEX(00007361)
    ws['A1'].font = Font(name='Arial',
                            sz=18,
                            color='00007361')

    # change alignment
    ws['A1'].alignment = Alignment(horizontal="left", vertical="top")

    # merge cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=12)

    # page orientation; This is not linked to the rest of the report. For easy printability it's all down
    #ws.page.PrintPageSetup()
    ws.page_setup.pageOrder = 'overThenDown'

    # TODO company logo
    

    # create header w/ page number or footer w/ text
    # ws.oddHeader.right.text = "Test bitch"
    # ws.oddHeader.right.size = 14
    # ws.oddHeader.right.font = 'Tahoma,Bold'
    # ws.oddHeader.right.color = 'CC3366'
    ws.oddFooter.right.text = "&[Page]/&N\n\n"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.font = 'Arial'

    ws.sheet_view.view = 'pageLayout'

    # create print area based on the number of lines the TOC has
    rowsTOC = 48 * pages
    ws.print_area = 'A1:L' + str(rowsTOC)
    
    # print chapters and paragraphs to the file
    line = 5
    for key, value in ChapAndPar.items():
        ws.cell(line, 1).value = key.split(" ", 1)[0]
        ws.cell(line, 1).font = Font(name='Arial',
                                    sz=12,
                                    color='00000000',
                                    bold=True)
        ws.cell(line, 1).alignment = Alignment(horizontal="left", vertical="bottom")

        ws.cell(line, 2).value = key.split(" ", 1)[1].lstrip().split("_", 1)[0]
        ws.cell(line, 2).font = Font(name='Arial',
                                    sz=12,
                                    color='00000000',
                                    bold=True)
        ws.cell(line, 2).alignment = Alignment(horizontal="left", vertical="bottom")

        ws.cell(line, 12).value = key.split(" ", 1)[1].lstrip().split("_", 1)[1]
        ws.cell(line, 12).font = Font(name='Arial',
                                    sz=12,
                                    color='00000000',
                                    bold=True)
        ws.cell(line, 12).alignment = Alignment(horizontal="right", vertical="bottom")
        line += 1

        for par in value:
            ws.cell(line, 1).value = par.split(" ", 1)[0]
            ws.cell(line, 1).font = Font(name='Arial',
                                        sz=10,
                                        color='00000000',
                                        bold=True)
            ws.cell(line, 1).alignment = Alignment(horizontal="left", vertical="bottom")

            ws.cell(line, 2).value = par.split(" ", 1)[1].lstrip().split("_", 1)[0]
            ws.cell(line, 2).font = Font(name='Arial',
                                        sz=10,
                                        color='00000000',
                                        bold=False)
            ws.cell(line, 2).alignment = Alignment(horizontal="left", vertical="bottom")

            ws.cell(line, 12).value = par.split(" ", 1)[1].lstrip().split("_", 1)[1]
            ws.cell(line, 12).font = Font(name='Arial',
                                        sz=10,
                                        color='00000000',
                                        bold=False)
            ws.cell(line, 12).alignment = Alignment(horizontal="right", vertical="bottom")
            line += 1
        ws.cell(line, 1).value = " "
        line += 1

    # close and save file with full table of content
    wb.save(path)
    wb.close()
